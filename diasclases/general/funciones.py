#!/usr/bin/env python
# -*- encoding: utf-8 -*-
from django.shortcuts import render, render_to_response
from django.template import RequestContext
from django.contrib.auth.models import Group, User
from django.http import HttpResponseRedirect, HttpResponse
from django.core.urlresolvers import reverse
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import permission_required, login_required
from diasclases.general.forms import *
from diasclases.general.models import *
from diasclases.monitoreo.models import *
import json
now = datetime.now()
diasemana = {'MONDAY':'Lunes','TUESDAY':'Martes','WEDNESDAY':'Miércoles',
             'THURSDAY':'Jueves','FRIDAY':'Viernes','SATURDAY':'Sábado',
             'SUNDAY':'Domingo'}
from datetime import datetime, timedelta
from django.db import transaction, IntegrityError
from openpyxl.workbook import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style, Color, Font, Alignment
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation, ValidationType


def my_range(start, end, step):
	while start <= end:
		yield start
		start += step

def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    data=[]
    desc = cursor.description
    for row in cursor.fetchall():
	    data.append({
	        'centro': row[0],
	        'total_si': row[1],
	        'total_no': row[2]
	    })
    return data

def validar_excel(archivo_excel):
	#archivo_excel = request.FILES['excel']
	#print "excel "+ str(archivo_excel)
	listaError=[]
	num_error=1
	try: 
		wb = load_workbook(filename = archivo_excel)
		#print wb.get_sheet_by_name(wb.get_sheet_names()[0])
		sheet_ranges = wb.get_sheet_by_name(wb.get_sheet_names()[0])
		#print sheet_ranges['A4'].value # D18
		razones1=[row[0] for row in razones.objects.all().values_list('id')]
		row=5
		
		while sheet_ranges['D%s'%(row)].value is not None:
			#print sheet_ranges['D%s'%(row)].value
			# Validar Nulos
			vacias=0
			for x in xrange(1,10):
				#print sheet_ranges['%s%s'%(get_column_letter(x),row)].value
				if sheet_ranges['%s%s'%(get_column_letter(x),row)].value is None:
					vacias=vacias+1
			if vacias > 0:
				listaError.append({
					'numero': num_error,
					'descripcion': 'Una o varias de las columnas (rango A-J) de la fila '+str(row)+' <b class="alert round label">estan vacías.</b> Por favor complete los datos e intentelo de nuevo.'
				})
				num_error=num_error+1
			# Validar que el centro exista y tenga voluntario
			if not voluntario.objects.filter(centros__codigo=sheet_ranges['D%s'%(row)].value.split('-')[0].strip(), centros__en_funcionamiento=True):
				#print voluntario.objects.filter(centros__codigo=sheet_ranges['D%s'%(row)].value, centros__anio=periodo.objects.get(activo=True).anio, centros__en_funcionamiento=True)
				listaError.append({
					'numero': num_error,
					'descripcion': 'El centro educativo ('+str(sheet_ranges['D%s'%(row)].value.split('-')[0].strip())+') de la fila '+str(row)+' <b class="alert round label"> no tiene un voluntario asignado.</b> Por favor asigne uno e intentelo de nuevo.'
				})
				num_error=num_error+1

			# Validar que el voluntario exista
			if not voluntario.objects.filter(identidad=sheet_ranges['G%s'%(row)].value.split('-')[0].strip(), activo=True):
				#print voluntario.objects.filter(centros__codigo=sheet_ranges['D%s'%(row)].value, centros__anio=periodo.objects.get(activo=True).anio, centros__en_funcionamiento=True)
				listaError.append({
					'numero': num_error,
					'descripcion': 'El voluntario del centro educativo ('+str(sheet_ranges['D%s'%(row)].value.split('-')[0].strip())+') de la fila '+str(row)+' <b class="alert round label">no esta registrado ó esta inactivo.</b>'
				})
				num_error=num_error+1

			# Validar que exista la combinacion centro voluntario
			if not voluntario.objects.filter(identidad=sheet_ranges['G%s'%(row)].value.split('-')[0].strip(), activo=True,centros__codigo=sheet_ranges['D%s'%(row)].value.split('-')[0].strip(), centros__en_funcionamiento=True):
				#print voluntario.objects.filter(centros__codigo=sheet_ranges['D%s'%(row)].value, centros__anio=periodo.objects.get(activo=True).anio, centros__en_funcionamiento=True)
				listaError.append({
					'numero': num_error,
					'descripcion': 'El voluntario de la fila '+str(row)+' <b class="alert round label"> NO es el encargado del centro</b> ('+str(sheet_ranges['D%s'%(row)].value.split('-')[0].strip())+').'
				})
				num_error=num_error+1

			# Validar el rango de fechas de la semana
			if not calendario_academico.objects.filter(fecha_inicio=datetime.strptime(sheet_ranges['I%s'%(row)].value, '%d-%m-%Y').date(), fecha_fin=datetime.strptime(sheet_ranges['J%s'%(row)].value, '%d-%m-%Y').date()):
				listaError.append({
					'numero': num_error,
					'descripcion': 'El centro educativo ('+str(sheet_ranges['D%s'%(row)].value).split('-')[0].strip()+') de la fila '+str(row)+' tiene una <b class="alert round label">fecha de inicio ó fecha fin que no estan registradas</b>. Por favor evite modificar estas columnas.'
				})
				num_error=num_error+1
				# Validar jornada
			if not jornada.objects.filter(pk=sheet_ranges['H%s'%(row)].value):
				listaError.append({
					'numero': num_error,
					'descripcion': 'El centro educativo ('+str(sheet_ranges['D%s'%(row)].value.split('-')[0].strip())+') de la fila '+str(row)+' tiene una <b class="alert round label">jornada no permitida</b>. Por faver revise los códigos de las jornadas.'
				})
				num_error=num_error+1
			# Validar registro de dias
			for x in my_range(11, 21, 2):
				if sheet_ranges['%s4'%(get_column_letter(x))].value is not None and sheet_ranges['%s%s'%(get_column_letter(x),row)].value is None:
					listaError.append({
						'numero': num_error,
						'descripcion': 'El día <b class="alert round label">'+str(sheet_ranges['%s4'%(get_column_letter(x))].value.encode('utf-8'))+'</b> del centro educativo ('+str(sheet_ranges['D%s'%(row)].value.split('-')[0].strip())+') de la fila '+str(row)+' no se a reportado.'
					})
					num_error=num_error+1

				if sheet_ranges['%s4'%(get_column_letter(x))].value is not None and sheet_ranges['%s%s'%(get_column_letter(x),row)].value.upper() not in ('S', 'N', None):
					listaError.append({
						'numero': num_error,
						'descripcion': 'El día <b>'+str(sheet_ranges['%s4'%(get_column_letter(x))].value.encode('utf-8'))+'</b> del centro educativo ('+str(sheet_ranges['D%s'%(row)].value.split('-')[0].strip())+') de la fila '+str(row)+' tiene un valor no permitido.'
					})
					num_error=num_error+1

				if sheet_ranges['%s4'%(get_column_letter(x))].value is not None and sheet_ranges['%s%s'%(get_column_letter(x),row)].value.upper() in ('N') and sheet_ranges['%s%s'%(get_column_letter(x+1),row)].value is None:
					listaError.append({
						'numero': num_error,
						'descripcion': 'Debe <b class="alert round label">especificar una razón</b> para el día <b>'+str(sheet_ranges['%s4'%(get_column_letter(x))].value.encode('utf-8'))+'</b> del centro educativo ('+str(sheet_ranges['D%s'%(row)].value.split('-')[0].strip())+') de la fila '+str(row)+'.'
					})
					num_error=num_error+1
				if sheet_ranges['%s4'%(get_column_letter(x))].value is not None and sheet_ranges['%s%s'%(get_column_letter(x),row)].value.upper() == 'N' and sheet_ranges['%s%s'%(get_column_letter(x+1),row)].value is not None and sheet_ranges['%s%s'%(get_column_letter(x+1),row)].value not in (razones1):
					print sheet_ranges['%s%s'%(get_column_letter(x+1),row)].value
					listaError.append({
						'numero': num_error,
						'descripcion': 'La <b class="alert round label">razón especificada para el día '+str(sheet_ranges['%s4'%(get_column_letter(x))].value.encode('utf-8'))+'</b> del centro educativo ('+str(sheet_ranges['D%s'%(row)].value.split('-')[0].strip())+') de la fila '+str(row)+' no está en la lista de opciones.'
					})
					num_error=num_error+1
			
			row=row+1
	except Exception, e:
		print "Error"
		print e;
		listaError.append({
			'numero': num_error,
			'descripcion': 'Una o más celdas tienen un <b class="alert round label">formato no permitido</b> por favor revise la información enviada y evite modificar las columnas de la A-J.'
		})
		num_error=num_error+1
	return listaError


@transaction.atomic
def guardar_excel(archivo_excel, usuario): 
		wb = load_workbook(filename = archivo_excel)
		#print wb.get_sheet_names()
		ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
		listaError=[]
		num_error=1
		duplicados=False
		periodos=periodo.objects.get(activo=True)
		#print sheet_ranges['A4'].value # D18
		row=5
		while ws['D%s'%(row)].value is not None:
			
			#if centro_voluntario.objects.filter(periodo=periodos.pk, centro=centro.pk, voluntario=voluntarios.pk, jornada=jornadas.pk):
			
			commit=None
			try:
					centro=centro_educativo.objects.get(codigo=ws['D%s'%(row)].value.split('-')[0].strip(),nivel=ws['D%s'%(row)].value.split(' - ')[1].strip())
					voluntarios=voluntario.objects.get(identidad=ws['G%s'%(row)].value.split('-')[0].strip())
					jornadas=jornada.objects.get(pk=ws['H%s'%(row)].value)
					sid = transaction.savepoint()
					insert_cv=centro_voluntario.objects.get_or_create(
						periodo=periodos,
						centro=centro,
						voluntario=voluntarios,
						jornada=jornadas
					)
					print "tipo_centro: "+ws['F%s'%(row)].value
					centro.tipo_docente=ws['F%s'%(row)].value
					centro.save()
					#insert_cv.save()
					calendario=calendario_academico.objects.get(fecha_inicio=datetime.strptime(ws['I%s'%(row)].value, '%d-%m-%Y').date(), fecha_fin=datetime.strptime(ws['J%s'%(row)].value, '%d-%m-%Y').date())
					cols=11
					for x in range(0,6):
						if duplicados:
							break
						fecha=calendario.fecha_inicio+timedelta(days=x)
						if fecha.strftime("%A") != 'Sunday':
							if centro_semana.objects.filter(centro=insert_cv[0].pk, semana=calendario.pk, fecha=fecha):
								listaError.append({
									'numero': num_error,
									'descripcion': str('La combinación Semana-Jornada que intenta ingresar para estos centros educativos <b class="alert round label">ya se ha almacenado</b> en la base de datos.')
								})
								num_error += 1
								transaction.savepoint_rollback(sid)
								duplicados=True
								break
							else:
								if ws['%s%s'%(get_column_letter(cols),row)].value.upper() == 'S':
									hubo_clases=True
									razon=None
								else:
									hubo_clases=False
									print get_column_letter(cols+1),row, ws['%s%s'%(get_column_letter(cols+1),row)].value
									razon=razones.objects.get(pk=ws['%s%s'%(get_column_letter(cols+1),row)].value)
								insert_cs=centro_semana(
									centro=insert_cv[0],
									semana=calendario,
									fecha=fecha,
									hubo_clases=hubo_clases,
									razon=razon,
									usuario_creador = User.objects.get(pk=usuario),
									fecha_creacion = datetime.now(),
									usuario_modificador = User.objects.get(pk=usuario),
									fecha_modificacion = datetime.now()
								)
								insert_cs.save()
						else:
							break
						cols += 2
					transaction.savepoint_commit(sid)
			except IntegrityError, e:
				print e
				listaError.append({
					'numero': num_error,
					'descripcion': str('La combinación Semana-Jornada que intenta ingresar para estos centros educativos <b class="alert round label">ya se ha almacenado</b> en la base de datos.')
				})
				num_error += 1
				transaction.savepoint_rollback(sid)
			except Exception, e:
				print e
				listaError.append({
					'numero': num_error,
					'descripcion': str("Error al cargar el archivo de excel.")
				})
				num_error += 1
				transaction.savepoint_rollback(sid)

			row += 1
		return listaError