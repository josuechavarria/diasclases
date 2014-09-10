#!/usr/bin/env python
# -*- encoding: utf-8 -*-
from django.shortcuts import render, render_to_response
from django.template import RequestContext
from django.contrib.auth.models import Group, User
from django.http import HttpResponseRedirect, HttpResponse
from django.core.urlresolvers import reverse
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import permission_required, login_required
from django.conf.global_settings import PASSWORD_HASHERS as default_hashers
from django.contrib.auth.hashers import (is_password_usable, 
    check_password, make_password, PBKDF2PasswordHasher, load_hashers,
    PBKDF2SHA1PasswordHasher, get_hasher)
from django.core.mail import EmailMultiAlternatives
from django.db.models import Q
from diasclases.general.forms import *
from diasclases.general.models import *
from diasclases.monitoreo.models import *

import json
from django.conf import settings
now = datetime.now()
diasemana = {'MONDAY':'Lunes','TUESDAY':'Martes','WEDNESDAY':'Miércoles',
             'THURSDAY':'Jueves','FRIDAY':'Viernes','SATURDAY':'Sábado',
             'SUNDAY':'Domingo'}
from django.db.models import Count, Sum, Avg
from datetime import datetime, timedelta
from django.db import transaction, IntegrityError
from openpyxl.workbook import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.styles import Style, Color, Font, Alignment, NumberFormat
from openpyxl import load_workbook
from openpyxl.datavalidation import DataValidation, ValidationType
from django.utils.encoding import smart_str
from django.db import connections
cursor = connections['default'].cursor()
from diasclases.general.funciones import *


def view_login(request):
	mensaje=""
	if request.user.is_authenticated():
		return HttpResponseRedirect(reverse('vista_inicio')) #redirecciona a la raiz
	else:
		if request.method=="POST":
			form = FormLogin(request.POST)
			if form.is_valid():
				username=(form.cleaned_data['username']).strip()
				password=(form.cleaned_data['password']).strip()
				usuario=authenticate(username=username, password=password)
				if usuario is not None and usuario.is_active: #si el usuario no es nullo y esta activo
					login(request, usuario) #crea la sesion
					user = User.objects.get(id=request.user.id)
					#mensaje = 'El usuario <<  '+username + ' >> no tiene los permisos necesarios para ingresar al modulo o no existe en el sistema.'
					return HttpResponseRedirect(reverse('vista_inicio'))
				else:
					mensaje = 'El usuario <<  '+username + ' >> no tiene los permisos necesarios para ingresar al modulo o no existe en el sistema.'
					logout(request)
		form = FormLogin()
		ctx={'form':form, 'mensaje':mensaje}			
		return render_to_response('general/login.html', ctx, context_instance=RequestContext(request))


def view_logout(request):
	logout(request)
	return HttpResponseRedirect('/')

@login_required
def view_index(request):
	ctx={'data':'data' }
	#print centro_semana.objects.values('centro','total_si').extra(select={'total_si':"Sum(case when hubo_clases=True then 1 else 0 end)"}).annotate().query
	return render_to_response('general/inicio_monitoreo.html',ctx, context_instance=RequestContext(request))


@permission_required('general.add_voluntario', login_url='/inicio/')
@transaction.atomic
def view_nueva_persona(request):
	#recuperar datos del centro
		fecha = datetime.strptime('2014-08-09', '%Y-%m-%d').date()# - timedelta(days=5)
		print fecha.strftime("%A")
		print fecha
		if request.method == 'POST':
			print "salvando1"
			formulario = PersonaForm(request.POST)

			if formulario.is_valid():
				sid = transaction.savepoint()
				print "es valido"
				try: 
					persona = formulario.save(commit = False)
					persona.activo=True
					persona.primer_nombre=(formulario.cleaned_data['primer_nombre']).upper()
					persona.segundo_nombre=(formulario.cleaned_data['segundo_nombre']).upper()
					persona.primer_apellido=(formulario.cleaned_data['primer_apellido']).upper()
					persona.segundo_apellido=(formulario.cleaned_data['segundo_apellido']).upper()
					persona.usuario_creador = User.objects.get(pk=request.user.pk)
					persona.fecha_creacion = datetime.now()
					persona.usuario_modificador = User.objects.get(pk=request.user.pk)
					persona.fecha_modificacion = datetime.now()
					persona.save()
					formulario.save_m2m()
					transaction.savepoint_commit(sid)
					#retornar exito
					formulario = PersonaForm()
					ctx = {'formulario': formulario, 'exito':"si"}
				except IntegrityError, e:
					print e
					transaction.savepoint_rollback(sid)
					mensaje = 'Uno de los centros seleccionados ya tiene un voluntario asignado.'
					ctx = {'formulario': formulario, 'error':'cx', 'errores': mensaje}
				except Exception, e:
					transaction.savepoint_rollback(sid)
					mensaje = 'Ocurrió un error al guardar los datos por favor verifique la información digitada.'
					ctx = {'formulario': formulario, 'error':'cx', 'errores': mensaje}
			else:
				ctx = {'formulario': formulario, 'error':'cx', 'errores': str(formulario.errors)}
				print "invalido"
				print formulario.errors
		else:
			formulario = PersonaForm()
			ctx = {'formulario': formulario}
		return render_to_response('general/nueva-persona.html', ctx, context_instance=RequestContext(request))

@permission_required('general.add_persona', login_url='/inicio/')
def view_listar_centros(request):
	#recuperar datos del centro
		centrolist = centro_educativo.objects.all().order_by('codigo')
		ctx = {'centros': centrolist}
		return render_to_response('general/listar_centros.html', ctx, context_instance=RequestContext(request))


@login_required
def view_people_ajax_municipio(request):
	if request.method == 'GET':
		if request.GET.get('bandera') == 'd':
			info=municipio.objects.filter(departamento=request.GET.get('id')).order_by('codigo_municipio')
			data = [{'pk':row.pk,'codigo':row.codigo_municipio,'nombre':row.descripcion} for row in info]
		elif request.GET.get('bandera') == 'm':
			info=aldea.objects.filter(municipio=request.GET.get('id')).order_by('codigo_aldea')
			data = [{'pk':row.pk,'codigo':row.codigo_aldea,'nombre':row.descripcion} for row in info]
		elif request.GET.get('bandera') == 'cm':
			info=centro_educativo.objects.filter(municipio=request.GET.get('id')).order_by('codigo')
			data = [{'pk':row.pk,'codigo':row.codigo,'nombre':row.nombre, 'nivel':row.nivel} for row in info]
		elif request.GET.get('bandera') == 'ca':
			info=centro_educativo.objects.filter(aldea=request.GET.get('id')).order_by('codigo')
			data = [{'pk':row.pk,'codigo':row.codigo,'nombre':row.nombre, 'nivel':row.nivel} for row in info]
		return HttpResponse(json.dumps(data), content_type="application/json")
	else:
		return HttpResponse(0)

@login_required
def view_centro_ajax(request):
	if request.method == 'GET':
		centro=centro_educativo.objects.filter(pk=request.GET.get('id')).order_by('codigo')
		voluntarios=voluntario.objects.filter(centros__id=request.GET.get('id'), centros__en_funcionamiento=True, activo=True)
		data = [{'pk':row.pk,'codigo':row.codigo,'nombre':row.nombre, 'tipo_docente':row.tipo_docente} for row in centro]
		print centro
		print voluntarios
		data2 = [{'pk':voluntarios.pk,'identidad':voluntarios.identidad,'telefono':voluntarios.telefono ,'nombre':voluntarios.primer_nombre+' '+voluntarios.primer_apellido} for voluntarios in voluntarios]

		data3 = [{'centro':data,'voluntario':data2}]
		return HttpResponse(json.dumps(data3), content_type="application/json")
	else:
		return HttpResponse(0)

@login_required
@permission_required('monitoreo.puede_descargar_excel', login_url='/inicio/')
def view_generar_excel(request):
	if request.method == 'POST':
		formulario = DepartamentoForm(request.POST)
		if formulario.is_valid():
			print "es valido"
			print request.POST.get('municipio'), request.POST.get('aldea')
			mun=""
			ald=""
			wb = Workbook()

			ws = wb.active
			semana = calendario_academico.objects.get(id=request.POST.get('semana'))
			#print calendario_academico.objects.all().values_list('centro_semana__centro__centro__pk').query
			if request.POST.get('municipio') == '' and request.POST.get('aldea') == '':
				centrov = voluntario.objects.filter(centros__anio=periodo.objects.get(activo=True).anio, centros__departamento__pk=request.POST.get('departamento')).values_list('identidad','centros','primer_nombre', 'primer_apellido').distinct('identidad','centros')

			if request.POST.get('municipio') != '' and request.POST.get('aldea') == '':
				centrov = voluntario.objects.filter(centros__anio=periodo.objects.get(activo=True).anio, centros__departamento__pk=request.POST.get('departamento'), centros__municipio__pk=request.POST.get('municipio')).values_list('identidad','centros').distinct('identidad','centros')
				mun=municipio.objects.get(pk=request.POST.get('municipio')).codigo_municipio
				ald=""

			if request.POST.get('municipio') != '' and request.POST.get('aldea') != '':
				centrov = voluntario.objects.filter(centros__anio=periodo.objects.get(activo=True).anio, centros__departamento__pk=request.POST.get('departamento'), centros__municipio__pk=request.POST.get('municipio'), centros__aldea__pk=request.POST.get('aldea')).values_list('identidad','centros').distinct('identidad','centros')
				mun=municipio.objects.get(pk=request.POST.get('municipio')).codigo_municipio
				ald=aldea.objects.get(pk=request.POST.get('aldea')).codigo_aldea


			dest_filename = departamento.objects.get(pk=request.POST.get('departamento')).codigo_departamento+mun+'_'+ 'reporte_semanal'+'.xlsx'
			response = HttpResponse(content_type="application/ms-excel")
			response['Content-Disposition'] = 'attachment; filename='+dest_filename
			print centrov
			for c in centrov:
				print "hello"
			dv = DataValidation(ValidationType.LIST, formula1='"UNIDOCENTE,BIDOCENTE,MULTIDOCENTE"', allow_blank=True)
			dv.set_error_message('El valor de tipo de centro no esta permitido', 'Entrada no permitida')
			ws.add_data_validation(dv)
			ws.title = "Semana del " + str(semana.fecha_inicio.strftime("%d-%b-%Y"))
			
			ws['A1'] = "Sistema de monitoreo de días clases año " + str(periodo.objects.get(activo=True).anio)
			ws['A1'].style = Style(font=Font(name='Calibri',size=16,bold=True, color=Color(rgb=('2F75B5'))), alignment=Alignment(horizontal='center', vertical='bottom'))
			ws.merge_cells('A1:N1')

			ws['A2'] = "Reporte semanal del " + str(semana.fecha_inicio.strftime("%d-%b-%Y")) + " al " + str(semana.fecha_fin.strftime("%d-%b-%Y"))
			ws['A2'].style = Style(font=Font(name='Calibri',size=12,bold=True, color=Color(rgb=('FF000000'))), alignment=Alignment(horizontal='center', vertical='bottom'))
			ws.merge_cells('A2:N2')

			sty1=Style(font=Font(name='Calibri',size=11,bold=True, color=Color(rgb=('FF000000'))), alignment=Alignment(horizontal='left', vertical='bottom'))

			ws['A4'] = "Departamento"
			ws['B4'] = "Municipio"
			ws['C4'] = "Aldea"
			ws['D4'] = "Código CE"
			ws['E4'] = "Nombre CE"
			ws['F4'] = "Tipo CE"
			ws['G4'] = "Identidad Voluntario"
			ws['H4'] = "Jornada"
			ws['I4'] = "Fecha Inicio"
			ws['J4'] = "Fecha Fin"

			for x in xrange(1,23):
				ws['%s4'%(get_column_letter(x))].style =sty1

			col=11
			for x in range(0,6):
				fecha=semana.fecha_inicio+timedelta(days=x)
				if fecha.strftime("%A") != 'Sunday':
					ws['%s4'%(get_column_letter(col))] = diasemana[(semana.fecha_inicio+timedelta(days=x)).strftime("%A").upper()]
					ws['%s4'%(get_column_letter(col+1))] = "Razón"
					col=col+2
				else:
					break

			row=5
			for c in centrov:
				centro=centro_educativo.objects.get(pk=c[1])
				person=voluntario.objects.get(identidad=c[0])
				ws.cell('%s%s'%('A', row)).value = str(centro.departamento)
				ws.cell('%s%s'%('B', row)).value = str(centro.municipio)
				ws.cell('%s%s'%('C', row)).value = str(centro.aldea)
				ws.cell('%s%s'%('D', row)).value = str(centro.codigo)+ ' - ' +str(centro.nivel.encode('utf-8'))
				ws.cell('%s%s'%('E', row)).value = str(centro.nombre.encode('utf-8'))
				ws.cell('%s%s'%('F', row)).value = str(centro.tipo_docente)
				ws.cell('%s%s'%('G', row)).value = str(person.identidad)+ ' - ' + str(person.primer_nombre.encode('utf-8')) +' '+ str(person.primer_apellido.encode('utf-8'))
				ws.cell('%s%s'%('I', row)).value = str(semana.fecha_inicio.strftime("%d-%m-%Y"))
				ws.cell('%s%s'%('J', row)).value = str(semana.fecha_fin.strftime("%d-%m-%Y"))

				dv.add_cell(ws.cell('%s%s'%('F', row)))

				if ws.cell('%s%s'%('K', 4)).value is not None:
					ws.cell('%s%s'%('K', row)).value = 'S'
				if ws.cell('%s%s'%('M', 4)).value is not None:
					ws.cell('%s%s'%('M', row)).value = 'S'
				if ws.cell('%s%s'%('O', 4)).value is not None:
					ws.cell('%s%s'%('O', row)).value = 'S'
				if ws.cell('%s%s'%('Q', 4)).value is not None:
					ws.cell('%s%s'%('Q', row)).value = 'S'
				if ws.cell('%s%s'%('S', 4)).value is not None:
					ws.cell('%s%s'%('S', row)).value = 'S'
				if ws.cell('%s%s'%('U', 4)).value is not None:
					ws.cell('%s%s'%('U', row)).value = 'S'


				row=row+1
			# Agregando instrucciones
			ws = wb.create_sheet()

			ws.title = 'Instrucciones'

			ws['A1'] = "Utilice estos códigos para colocarlos en JORNADA Y EN RAZÓN respectivamente"
			ws['A1'].style = Style(font=Font(name='Calibri',size=13,bold=True, color=Color(rgb=('2F75B5'))), alignment=Alignment(horizontal='center', vertical='bottom'))
			ws.merge_cells('A1:K1')

			ws['A3'] = "JORNADAS"
			ws['A3'].style = Style(font=Font(name='Calibri',size=11,bold=True), alignment=Alignment(horizontal='center', vertical='bottom'))
			ws.merge_cells('A3:C3')
			rowy=4
			for j in jornada.objects.all().order_by('id'):
				ws['A%s'%(rowy)]=j.pk
				ws['B%s'%(rowy)]=j.descripcion
				rowy += 1
			rowy += 1
			ws['A%s'%(rowy)] = "RAZONES DE NO CLASES"
			ws['A%s'%(rowy)].style = Style(font=Font(name='Calibri',size=11,bold=True), alignment=Alignment(horizontal='center', vertical='bottom'))
			ws.merge_cells('A%s:C%s'%(rowy, rowy))
			rowy += 1
			for r in razones.objects.all().order_by('id'):
				ws['A%s'%(rowy)]=r.pk
				ws['B%s'%(rowy)]=r.descripcion
				rowy += 1

			wb.save(response)
			return response
	if request.user.is_superuser:
		formulario = DepartamentoForm(request.POST)
	else:
		filtro=persona.objects.filter(user=request.user.pk).values_list('departamentos_asignados__id')
		formulario = DepartamentoForm2(filtro, request.POST)
	return render_to_response('general/descargar_formato.html', {'formulario':formulario}, context_instance=RequestContext(request))

@login_required
def view_leer_excel(request):
	if request.method == 'POST':

		resultado=validar_excel(request.FILES['excel'])
		if len(resultado) > 0:
			return render_to_response('general/cargar_formato.html', {'mensaje': resultado} ,context_instance=RequestContext(request))
		else:
			resultado=guardar_excel(request.FILES['excel'], request.user.pk)
			if len(resultado) > 0:
				return render_to_response('general/cargar_formato.html', {'mensaje': resultado} ,context_instance=RequestContext(request))
			else:
				print "se guardo"
				return render_to_response('general/cargar_formato.html', {'completado': 'si', 'exito': 'si'} ,context_instance=RequestContext(request))
	return render_to_response('general/cargar_formato.html', context_instance=RequestContext(request))

@permission_required('general.puede_consultar_centros', login_url='/inicio/')
def view_consultar_centros(request):
	centro_sem=None
	new_list=[]
	if request.method == 'POST':
		formulario = ConsultaForm(request.POST)
		if formulario.is_valid():
			print "es valido"

			if request.POST.get('municipio') == '' and request.POST.get('aldea') == '':
				centro_sem = centro_semana.objects.filter(centro__centro__departamento__pk=request.POST.get('departamento'), centro__centro__anio=periodo.objects.get(activo=True).anio).distinct('centro')

			if request.POST.get('municipio') != '' and request.POST.get('aldea') == '':
				centro_sem = centro_semana.objects.filter(centro__centro__departamento__pk=request.POST.get('departamento'),centro__centro__municipio__pk=request.POST.get('municipio'), centro__centro__anio=periodo.objects.get(activo=True).anio).distinct('centro')

			if request.POST.get('municipio') != '' and request.POST.get('aldea') != '':
				centro_sem = centro_semana.objects.filter(centro__centro__departamento__pk=request.POST.get('departamento'),centro__centro__municipio__pk=request.POST.get('municipio'),centro__centro__aldea__pk=request.POST.get('aldea'), centro__centro__anio=periodo.objects.get(activo=True).anio).distinct('centro')

			
			for row in centro_sem:
				total_si=centro_semana.objects.filter(centro=row.centro, hubo_clases=True).count()
				total_no=centro_semana.objects.filter(centro=row.centro, hubo_clases=False).count()
				new_list.append({
					'centro': row,
					'total_si': total_si,
					'total_no': total_no
					})
		else:
			print formulario.errors

	if request.user.is_superuser:
		formulario = ConsultaForm(request.POST)
	else:
		filtro=persona.objects.filter(user=request.user.pk).values_list('departamentos_asignados__id')
		formulario = ConsultaForm2(filtro, request.POST)
	return render_to_response('general/consultar_centros.html', {'formulario': formulario, 'centro':new_list} , context_instance=RequestContext(request))

@permission_required('general.puede_consultar_detalle_centros', login_url='/inicio/')
def view_consultar_semanas(request, centro_voluntario_id):
	centro_sem = centro_semana.objects.filter(centro=centro_voluntario_id).order_by('semana', 'fecha')
	centro_info = centro_voluntario.objects.get(pk=centro_voluntario_id)
	return render_to_response('general/consultar_semanas_centro.html', {'centro':centro_sem, 'centro_info': centro_info} , context_instance=RequestContext(request))


@permission_required('general.puede_consultar_voluntarios', login_url='/inicio/')
def view_consultar_voluntarios(request):
	centro_vol=None
	if request.method == 'POST':
		formulario = ConsultaForm(request.POST)
		if formulario.is_valid():
			print "es valido"
			if request.POST.get('municipio') == '' and request.POST.get('aldea') == '':
				centro_vol = voluntario.objects.filter(departamento__pk=request.POST.get('departamento'), activo=True)

			if request.POST.get('municipio') != '' and request.POST.get('aldea') == '':
				centro_vol = voluntario.objects.filter(departamento__pk=request.POST.get('departamento'), municipio__pk=request.POST.get('municipio'), activo=True)

			if request.POST.get('municipio') != '' and request.POST.get('aldea') != '':
				centro_vol = voluntario.objects.filter(departamento__pk=request.POST.get('departamento'), municipio__pk=request.POST.get('municipio'), aldea__pk=request.POST.get('aldea'), activo=True)
		else:
			print formulario.errors
	if request.user.is_superuser:
		formulario = ConsultaForm(request.POST)
	else:
		filtro=persona.objects.filter(user=request.user.pk).values_list('departamentos_asignados__id')
		formulario = ConsultaForm2(filtro, request.POST)
	return render_to_response('general/consultar_voluntarios.html', {'formulario': formulario, 'centro':centro_vol} , context_instance=RequestContext(request))



@permission_required('general.change_persona', login_url='/inicio/')
@transaction.atomic
def view_nuevo_facilitador(request):
	#recuperar datos del centro
		fecha = datetime.strptime('2014-08-09', '%Y-%m-%d').date()# - timedelta(days=5)
		print fecha.strftime("%A")
		print fecha
		if request.method == 'POST':
			print "salvando1"
			formulario = Persona2Form(request.POST)

			if formulario.is_valid():
				sid = transaction.savepoint()
				print "es valido"
				try: 
					persona = formulario.save(commit = False)
					persona.activo=True
					persona.primer_nombre=(formulario.cleaned_data['primer_nombre']).upper()
					persona.segundo_nombre=(formulario.cleaned_data['segundo_nombre']).upper()
					persona.primer_apellido=(formulario.cleaned_data['primer_apellido']).upper()
					persona.segundo_apellido=(formulario.cleaned_data['segundo_apellido']).upper()
					persona.usuario_creador = User.objects.get(pk=request.user.pk)
					persona.fecha_creacion = datetime.now()
					persona.usuario_modificador = User.objects.get(pk=request.user.pk)
					persona.fecha_modificacion = datetime.now()

					#Crear el usuario de la persona
					clave = User.objects.make_random_password(length=8, allowed_chars='12345678#wertyupasdfghjkzxcvbnm')
					usuario = formulario.cleaned_data['identidad']
					staff=False
					superuser=False
					print request.POST.get('tipo_persona')
					if request.POST.get('tipo_persona') == '2':
						print 'facilitador'
						grupo = Group.objects.get(id=1)
					elif request.POST.get('tipo_persona') == '3':
						grupo = Group.objects.get(id=2)
					elif request.POST.get('tipo_persona') == '4':
						grupo = Group.objects.get(id=3)
						superuser=True
					user = User.objects.create_user(
						username=usuario, 
						email=formulario.cleaned_data['email'], 
						password=clave,
						first_name=(formulario.cleaned_data['primer_nombre']).upper(),
						last_name=(formulario.cleaned_data['primer_apellido']).upper(),
						)
					user.is_superuser=superuser
					user.is_staff=staff
					user.groups.add(grupo)
					user.save()
					persona.user=user
					persona.save()
					formulario.save_m2m()
					html_content='<h2>Bienvenido '+ formulario.cleaned_data['primer_nombre'] + ' ' + formulario.cleaned_data['primer_apellido'] + '</h2>'# se concatena con el texto de un formulario
					html_content += '<p style="font-size: 14px;"> <h4>Su usuario y contrasena para ingresar a nuestro sistema es:</h4> </p>'
					html_content += '<br>	Usuario:	<b>'+usuario+'</b><br>	Contrasena: <b>'+clave+'</b><br/>Privilegios de: <b>'+grupo.name+'</b><br/><br/> Att.<p><h3>El equipo de Transformemos Honduras</h3></p><img src="http://ajs-us.org/sites/default/files/styles/medium/public/projects/Transformemos_Honduras_logo.jpg?itok=jjXJYJiI">'
					#send_mail('Asunto del correo', cuerpo, 'josuechavarria89@gmail.com', [email])
					msg=EmailMultiAlternatives('Datos de acceso sistema de monitoreo de dias clases',html_content,'recursos.humanos.seduc.hn@gmail.com',[formulario.cleaned_data['email']])
					msg.attach_alternative(html_content,'text/html')
					#msg.attach(settings.MEDIA_ROOT+'/estaticos/img/logo2.jpg', 'img_data', 'image/jpg')
					#msg.attach_file(settings.MEDIA_ROOT+'/estaticos/img/logo2.jpg')
					try:
						msg.send()
						noenvio = '1'
						mensaje = u"Un correo electronico ha sido enviado a " + formulario.cleaned_data['email'] + " favor revisar para obtener sus datos de acceso"
						transaction.savepoint_commit(sid)
						#transaction.savepoint_rollback(sid)
						#retornar exito
						formulario = Persona2Form()
						ctx = {'formulario': formulario, 'exito':"si"}
					except Exception, e:
						noenvio = '0'
						print e
						transaction.savepoint_rollback(sid)
						mensaje = 'Ocurrió un error al enviar el EMAIL porfavor verifique que el servidor de correos este funcionando correctamente.'
						ctx = {'formulario': formulario, 'error':'cx', 'errores': mensaje}

				except IntegrityError, e:
					print e
					transaction.savepoint_rollback(sid)
					mensaje = 'Uno de los centros seleccionados ya tiene un voluntario asignado.'
					ctx = {'formulario': formulario, 'error':'cx', 'errores': mensaje}
				except Exception, e:
					print e
					transaction.savepoint_rollback(sid)
					mensaje = 'Ocurrió un error al guardar los datos por favor verifique la información digitada.'
					ctx = {'formulario': formulario, 'error':'cx', 'errores': mensaje}
			else:
				ctx = {'formulario': formulario, 'error':'cx', 'errores': str(formulario.errors)}
				print "invalido"
				print formulario.errors
		else:
			formulario = Persona2Form()
			ctx = {'formulario': formulario}
		return render_to_response('general/nueva-persona-facilitador.html', ctx, context_instance=RequestContext(request))


@permission_required('general.change_persona', login_url='/inicio/')
@transaction.atomic
def view_editar_facilitador(request, persona_id):
	#recuperar datos del centro
		personaObject = persona.objects.get(pk=persona_id)

		if request.method == 'POST':
			print "salvando1"
			formulario = Persona2EditForm(request.POST, instance=personaObject)

			if formulario.is_valid():
				sid = transaction.savepoint()
				print "es valido"
				try: 
					person = formulario.save(commit = False)
					#person.activo=True
					person.primer_nombre=(formulario.cleaned_data['primer_nombre']).upper()
					person.segundo_nombre=(formulario.cleaned_data['segundo_nombre']).upper()
					person.primer_apellido=(formulario.cleaned_data['primer_apellido']).upper()
					person.segundo_apellido=(formulario.cleaned_data['segundo_apellido']).upper()
					person.usuario_modificador = User.objects.get(pk=request.user.pk)
					person.fecha_modificacion = datetime.now()

					person.save()
					formulario.save_m2m()

					staff=False
					superuser=False
					usuario=User.objects.get(pk=person.user.pk)
					print request.POST.get('tipo_persona')
					if request.POST.get('tipo_persona') == '2':
						print 'facilitador'
						grupo = Group.objects.get(id=1)
					elif request.POST.get('tipo_persona') == '3':
						grupo = Group.objects.get(id=2)
					elif request.POST.get('tipo_persona') == '4':
						grupo = Group.objects.get(id=3)
						superuser=True

					usuario.is_superuser=superuser
					usuario.is_staff=staff
					usuario.first_name=(formulario.cleaned_data['primer_nombre']).upper()
					usuario.last_name=(formulario.cleaned_data['primer_apellido']).upper()
					usuario.save()
					user_group = User.groups.through.objects.get(user=usuario)
					user_group.group = grupo
					user_group.save()
					
					
					transaction.savepoint_commit(sid)
					#retornar exito
					formulario = Persona2EditForm(instance=personaObject)
					ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea'), 'exito':"si"}

				except IntegrityError, e:
					print e
					transaction.savepoint_rollback(sid)
					mensaje = 'Error de integridad de base de datos.'
					ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea'), 'error':'cx', 'errores': mensaje}
				except Exception, e:
					print e
					transaction.savepoint_rollback(sid)
					mensaje = 'Ocurrió un error al guardar los datos por favor verifique la información digitada.'
					ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea'), 'error':'cx', 'errores': mensaje}
			else:
				ctx = {'formulario': formulario, 'error':'cx', 'errores': str(formulario.errors)}
				print "invalido"
				print formulario.errors
		else:
			formulario = Persona2EditForm(instance=personaObject)
			ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea')}
		return render_to_response('general/editar-persona-facilitador.html', ctx, context_instance=RequestContext(request))

@permission_required('general.puede_buscar_usuarios', login_url='/inicio/')
@transaction.atomic
def view_buscar_usuarios(request):
	personaObject = None
	if request.method == 'POST':
		if request.POST.get('identidad').strip() != '' and (request.POST.get('primer_nombre').strip() == '' and request.POST.get('primer_apellido').strip() == ''):
			personaObject = persona.objects.filter(Q(identidad__contains=request.POST.get('identidad')), activo=True)
		if request.POST.get('primer_nombre').strip() != '' and request.POST.get('primer_apellido').strip() == '':
			print "entro 1"
			personaObject = persona.objects.filter(Q(primer_nombre__contains=request.POST.get('primer_nombre').upper()), activo=True)
		if request.POST.get('primer_nombre').strip() == '' and request.POST.get('primer_apellido').strip() != '':
			print "entro 1"
			personaObject = persona.objects.filter(Q(primer_apellido__contains=request.POST.get('primer_apellido').upper()), activo=True)
		if request.POST.get('primer_nombre').strip() != '' and request.POST.get('primer_apellido').strip() != '':
			personaObject = persona.objects.filter(Q(primer_nombre__contains=request.POST.get('primer_nombre').upper()),Q(primer_apellido__contains=request.POST.get('primer_apellido').upper()), activo=True)
	ctx = {'personas': personaObject}
	return render_to_response('general/buscar-usuario.html', ctx, context_instance=RequestContext(request))

@permission_required('general.puede_buscar_usuarios', login_url='/inicio/')
@transaction.atomic
def view_buscar_usuarios_restablecer(request):
	personaObject = None
	if request.method == 'POST':
		if request.POST.get('identidad').strip() != '' and (request.POST.get('primer_nombre').strip() == '' and request.POST.get('primer_apellido').strip() == ''):
			personaObject = persona.objects.filter(Q(identidad__contains=request.POST.get('identidad')), activo=True)
		if request.POST.get('primer_nombre').strip() != '' and request.POST.get('primer_apellido').strip() == '':
			print "entro 1"
			personaObject = persona.objects.filter(Q(primer_nombre__contains=request.POST.get('primer_nombre').upper()), activo=True)
		if request.POST.get('primer_nombre').strip() == '' and request.POST.get('primer_apellido').strip() != '':
			print "entro 1"
			personaObject = persona.objects.filter(Q(primer_apellido__contains=request.POST.get('primer_apellido').upper()), activo=True)
		if request.POST.get('primer_nombre').strip() != '' and request.POST.get('primer_apellido').strip() != '':
			personaObject = persona.objects.filter(Q(primer_nombre__contains=request.POST.get('primer_nombre').upper()),Q(primer_apellido__contains=request.POST.get('primer_apellido').upper()), activo=True)
	ctx = {'personas': personaObject}
	return render_to_response('general/buscar-usuario-restablecer.html', ctx, context_instance=RequestContext(request))

@permission_required('auth.change_user', login_url='/inicio/')
@transaction.atomic
def view_restablecer_clave(request, persona_id):
	personaObject=persona.objects.get(pk=persona_id)
	ctx={'persona': personaObject}
	if request.method == 'POST':
		
		try:
			if request.POST.get('email').strip() == '':
				sid = transaction.savepoint()
				user = User.objects.get(pk=personaObject.user.pk)
				user.set_password(request.POST.get('nueva_clave'))
				user.save()
			else:
				sid = transaction.savepoint()
				user = User.objects.get(pk=personaObject.user.pk)
				user.set_password(request.POST.get('nueva_clave'))
				user.save()
				user_group = User.groups.through.objects.get(user=user)
				
				html_content='<h2>Cambio de contrasena finalizado!</h2>'# se concatena con el texto de un formulario
				html_content += '<p style="font-size: 14px;"> <h4>Su nuevo usuario y contrasena para ingresar a nuestro sistema es:</h4> </p>'
				html_content += '<br>	Usuario:	<b>'+user.username+'</b><br>	Contrasena: <b>'+request.POST.get('nueva_clave')+'</b><br/>Privilegios de: <b>'+user_group.group.name+'</b><br/><br/> Att.<p><h3>El equipo de Transformemos Honduras</h3></p><img src="http://ajs-us.org/sites/default/files/styles/medium/public/projects/Transformemos_Honduras_logo.jpg?itok=jjXJYJiI">'
				#send_mail('Asunto del correo', cuerpo, 'josuechavarria89@gmail.com', [email])
				msg=EmailMultiAlternatives('Solicitud cambio de contrasena sistema de monitoreo de dias clases',html_content,'recursos.humanos.seduc.hn@gmail.com',[request.POST.get('email')])
				msg.attach_alternative(html_content,'text/html')
				try:
					msg.send()
					noenvio = '1'
					mensaje = u"Un correo electronico ha sido enviado a " + request.POST.get('email') + " favor revisar para obtener sus datos de acceso"
					#retornar exito
					ctx = {'persona': personaObject, 'exito':"si"}
				except Exception, e:
					noenvio = '0'
					print e
					transaction.savepoint_rollback(sid)
					mensaje = 'Ocurrió un error al enviar el EMAIL porfavor verifique que el servidor de correos este funcionando correctamente.'
					ctx = {'persona': personaObject, 'error':'cx', 'errores': mensaje}

			transaction.savepoint_commit(sid)
			ctx = {'persona': personaObject, 'nueva_clave': request.POST.get('nueva_clave'), 'email': request.POST.get('email'), 'exito':"si"}
		except Exception, e:
			print e
			transaction.savepoint_rollback(sid)
			mensaje = 'Ocurrió un error al guardar los datos por favor verifique la información digitada.'
			ctx = {'persona': personaObject, 'error':'cx', 'errores': mensaje}

	return render_to_response('general/usuario-restablecer-clave.html', ctx, context_instance=RequestContext(request))

@permission_required('auth.change_user', login_url='/inicio/')
@transaction.atomic
def view_ajax_generar(request):
	if request.method == 'GET':
		clave = User.objects.make_random_password(length=8, allowed_chars='12345678#wertyupasdfghjkzxcvbnm')
		data = [{'clave':clave}]
		return HttpResponse(json.dumps(data), content_type="application/json")
	else:
		return HttpResponse(0)

@permission_required('general.puede_buscar_usuarios', login_url='/inicio/')
@transaction.atomic
def view_buscar_usuarios_deshabilitar(request):
	personaObject = None
	if request.method == 'POST':
		if request.POST.get('identidad').strip() != '' and (request.POST.get('primer_nombre').strip() == '' and request.POST.get('primer_apellido').strip() == ''):
			personaObject = persona.objects.filter(Q(identidad__contains=request.POST.get('identidad')))
		if request.POST.get('primer_nombre').strip() != '' and request.POST.get('primer_apellido').strip() == '':
			print "entro 1"
			personaObject = persona.objects.filter(Q(primer_nombre__contains=request.POST.get('primer_nombre').upper()))
		if request.POST.get('primer_nombre').strip() == '' and request.POST.get('primer_apellido').strip() != '':
			print "entro 1"
			personaObject = persona.objects.filter(Q(primer_apellido__contains=request.POST.get('primer_apellido').upper()))
		if request.POST.get('primer_nombre').strip() != '' and request.POST.get('primer_apellido').strip() != '':
			personaObject = persona.objects.filter(Q(primer_nombre__contains=request.POST.get('primer_nombre').upper()),Q(primer_apellido__contains=request.POST.get('primer_apellido').upper()))
	ctx = {'personas': personaObject}
	return render_to_response('general/buscar-usuario-deshabilitar.html', ctx, context_instance=RequestContext(request))

@permission_required('auth.change_user', login_url='/inicio/')
@transaction.atomic
def view_usuario_deshabilitar(request, persona_id):
	sid = transaction.savepoint()
	try:
		personaObject=persona.objects.get(pk=persona_id)
		user=User.objects.get(pk=personaObject.user.pk)
		user.is_active=False
		personaObject.activo=False
		user.save()
		personaObject.save()
		transaction.savepoint_commit(sid)
		ctx = {'persona': personaObject, 'exito':"si", 'mensaje': 'Usuario deshabilitado correctamente.'}
		return HttpResponse(1)
	except Exception, e:
		print e
		transaction.savepoint_rollback(sid)
		mensaje = 'Ocurrió un error al deshabilitar el usuario.'
		ctx = {'persona': personaObject, 'error':'cx', 'errores': mensaje}
	return HttpResponse(0)

@permission_required('auth.change_user', login_url='/inicio/')
@transaction.atomic
def view_usuario_habilitar(request, persona_id):
	sid = transaction.savepoint()
	try:
		personaObject=persona.objects.get(pk=persona_id)
		user=User.objects.get(pk=personaObject.user.pk)
		user.is_active=True
		personaObject.activo=True
		user.save()
		personaObject.save()
		transaction.savepoint_commit(sid)
		ctx = {'persona': personaObject, 'exito':"si", 'mensaje': 'Usuario deshabilitado correctamente.'}
		return HttpResponse(1)
	except Exception, e:
		print e
		transaction.savepoint_rollback(sid)
		mensaje = 'Ocurrió un error al deshabilitar el usuario.'
		ctx = {'persona': personaObject, 'error':'cx', 'errores': mensaje}
	return HttpResponse(0)


@permission_required('general.puede_consultar_voluntarios', login_url='/inicio/')
@transaction.atomic
def view_buscar_voluntarios(request):
	personaObject = None
	if request.method == 'POST':
		if request.POST.get('identidad').strip() != '' and (request.POST.get('primer_nombre').strip() == '' and request.POST.get('primer_apellido').strip() == ''):
			personaObject = voluntario.objects.filter(Q(identidad__contains=request.POST.get('identidad')))
		if request.POST.get('primer_nombre').strip() != '' and request.POST.get('primer_apellido').strip() == '':
			personaObject = voluntario.objects.filter(Q(primer_nombre__contains=request.POST.get('primer_nombre').upper()))
		if request.POST.get('primer_nombre').strip() == '' and request.POST.get('primer_apellido').strip() != '':
			personaObject = voluntario.objects.filter(Q(primer_apellido__contains=request.POST.get('primer_apellido').upper()))
		if request.POST.get('primer_nombre').strip() != '' and request.POST.get('primer_apellido').strip() != '':
			personaObject = voluntario.objects.filter(Q(primer_nombre__contains=request.POST.get('primer_nombre').upper()),Q(primer_apellido__contains=request.POST.get('primer_apellido').upper()))
	ctx = {'personas': personaObject}
	return render_to_response('general/buscar-voluntario.html', ctx, context_instance=RequestContext(request))

@permission_required('general.change_voluntario', login_url='/inicio/')
@transaction.atomic
def view_editar_voluntario(request, persona_id):
	#recuperar datos del centro
		personaObject = voluntario.objects.get(pk=persona_id)

		if request.method == 'POST':
			print "salvando1"
			formulario = PersonaEditForm(request.POST, instance=personaObject)

			if formulario.is_valid():
				sid = transaction.savepoint()
				print "es valido"
				try: 
					person = formulario.save(commit = False)
					if request.POST.get('activo') == 'False':
						activo=False
					else:
						activo=True
					#person.activo=True
					person.primer_nombre=(formulario.cleaned_data['primer_nombre']).upper()
					person.segundo_nombre=(formulario.cleaned_data['segundo_nombre']).upper()
					person.primer_apellido=(formulario.cleaned_data['primer_apellido']).upper()
					person.segundo_apellido=(formulario.cleaned_data['segundo_apellido']).upper()
					person.activo=activo
					person.usuario_modificador = User.objects.get(pk=request.user.pk)
					person.fecha_modificacion = datetime.now()

					person.save()
					formulario.save_m2m()
					transaction.savepoint_commit(sid)
					#retornar exito
					formulario = PersonaEditForm(instance=personaObject)
					ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea'),
			'centros':centro_educativo.objects.filter(aldea=personaObject.aldea.pk).order_by('codigo'), 'exito':"si"}

				except IntegrityError, e:
					print e
					transaction.savepoint_rollback(sid)
					mensaje = 'Error de integridad de base de datos.'
					ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea'), 'error':'cx', 'errores': mensaje}
				except Exception, e:
					print e
					transaction.savepoint_rollback(sid)
					mensaje = 'Ocurrió un error al guardar los datos por favor verifique la información digitada.'
					ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea'), 'error':'cx', 'errores': mensaje}
			else:
				ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea'), 'error':'cx', 'errores': str(formulario.errors)}
				print "invalido"
				print formulario.errors
		else:
			formulario = PersonaEditForm(instance=personaObject)
			ctx = {'formulario': formulario, 'persona': personaObject,
			'municipio': municipio.objects.filter(departamento__pk=personaObject.municipio.departamento.pk), 'aldea': aldea.objects.filter(municipio__pk=personaObject.municipio.pk).order_by('codigo_aldea'),
			'centros':centro_educativo.objects.filter(aldea=personaObject.aldea.pk).order_by('codigo')}
		return render_to_response('general/editar-persona-voluntario.html', ctx, context_instance=RequestContext(request))


@permission_required('general.add_calendario_academico', login_url='/inicio/')
@transaction.atomic
def view_generar_calendario(request):
	ctx={'x':'x'}
	if request.method == 'POST':
		diferencia=datetime.strptime(request.POST.get('fecha_fin_anio'), '%Y-%m-%d').date()-datetime.strptime(request.POST.get('fecha_inicio_anio'), '%Y-%m-%d').date()
		if datetime.strptime(request.POST.get('fecha_fin_anio'), '%Y-%m-%d').date() < datetime.strptime(request.POST.get('fecha_inicio_anio'), '%Y-%m-%d').date():
			mensaje = 'No se pudo generar el calendario. La fecha fin no puede ser menor que la fecha de inicio.'
			ctx = {'error':'cx', 'errores': mensaje}
			return render_to_response('general/generar-calendario.html', ctx ,context_instance=RequestContext(request))
		if diferencia.days < 200:
			mensaje = 'No se pudo generar el calendario. La cantidad de dias en el anio debe ser al menos 200 dias.'
			ctx = {'error':'cx', 'errores': mensaje}
			return render_to_response('general/generar-calendario.html', ctx ,context_instance=RequestContext(request))

		calendarioList=[]
		fechasExclude=[]
		fechasExclude2=[]
		for x in xrange(1,5):
			if request.POST.get('fecha_inicio_'+str(x)).strip() != '':
				fecha_inicio=datetime.strptime(request.POST.get('fecha_inicio_'+str(x)), '%Y-%m-%d').date()
				fecha_fin=datetime.strptime(request.POST.get('fecha_fin_'+str(x)), '%Y-%m-%d').date()
				fechasExclude.append({
					'fecha_inicio': fecha_inicio
				})
				print fecha_inicio, fecha_fin
				print (fecha_fin-fecha_inicio).days
				if (fecha_fin-fecha_inicio).days < 5 or (fecha_fin-fecha_inicio).days > 5 or fecha_fin < fecha_inicio:
					mensaje = ' En las semanas feriados. La fecha fin no puede ser menor que la fecha de inicio, la semana debe tener exactamente 6 dias.'
					ctx = {'error':'cx', 'errores': mensaje}
					return render_to_response('general/generar-calendario.html', ctx ,context_instance=RequestContext(request))
		
		fechasExclude2=[row['fecha_inicio'] for row in fechasExclude]
		#print fechasExclude2

		fecha_inicio=datetime.strptime(request.POST.get('fecha_inicio_anio'), '%Y-%m-%d').date()
		fecha_fin=None
		countWeek=1
		while fecha_fin != datetime.strptime(request.POST.get('fecha_fin_anio'), '%Y-%m-%d').date():
			for x in range(0,6):
				
				fecha_fin=fecha_inicio+timedelta(days=x)
				if fecha_fin != datetime.strptime(request.POST.get('fecha_fin_anio'), '%Y-%m-%d').date() and fecha_fin.strftime("%A") == 'Saturday' and fecha_inicio not in (fechasExclude2):
					calendarioList.append({
						'numero_semana' : countWeek,
						'fecha_inicio': fecha_inicio,
						'fecha_fin': fecha_fin
					})
					fecha_inicio=fecha_fin+timedelta(days=2)
					countWeek += 1
					#print calendarioList
					#raw_input("Pulsa una tecla para continuar...") 
				elif fecha_fin == datetime.strptime(request.POST.get('fecha_fin_anio'), '%Y-%m-%d').date():
					calendarioList.append({
						'numero_semana' : countWeek,
						'fecha_inicio': fecha_inicio,
						'fecha_fin': fecha_fin
					})
					#print calendarioList
					countWeek += 1
					#raw_input("Pulsa una tecla para continuar...")
					break
				elif fecha_fin != datetime.strptime(request.POST.get('fecha_fin_anio'), '%Y-%m-%d').date() and fecha_fin.strftime("%A") == 'Saturday' and fecha_inicio in (fechasExclude2):
					fecha_inicio=fecha_fin+timedelta(days=2)
					#break

		try:
			sid = transaction.savepoint()
			for row in calendarioList:
				if calendario_academico.objects.filter(fecha_fin=row['fecha_fin']):
					transaction.savepoint_rollback(sid)
					mensaje = 'No se pudo generar el calendario. El anio que esta ingresando ya existe.'
					ctx = {'error':'cx', 'errores': mensaje}
					return render_to_response('general/generar-calendario.html', ctx ,context_instance=RequestContext(request))
				else:
					insert=calendario_academico(
							numero_semana=row['numero_semana'],
							fecha_inicio=row['fecha_inicio'],
							fecha_fin=row['fecha_fin']
						)
					insert.save()
			transaction.savepoint_commit(sid)
			ctx = {'exito':'si'}
		except Exception, e:
			print e
			transaction.savepoint_rollback(sid)
			mensaje = 'No se pudo generar el calendario.'
			ctx = {'error':'cx', 'errores': mensaje}
		#print calendarioList
	return render_to_response('general/generar-calendario.html', ctx ,context_instance=RequestContext(request))

@permission_required('monitoreo.delete_calendario_academico', login_url='/inicio/')
def view_listar_calendario(request):
	ctx={}
	ctx={'calendario': calendario_academico.objects.filter(fecha_inicio__year=periodo.objects.get(activo=True).anio), 'anio':periodo.objects.get(activo=True).anio}
	return render_to_response('general/listar-calendario.html', ctx ,context_instance=RequestContext(request))

@permission_required('monitoreo.delete_calendario_academico', login_url='/inicio/')
def view_eliminar_semanacalendario(request):
	ctx={}
	ctx={'calendario': calendario_academico.objects.exclude(pk__in=centro_semana.objects.filter(fecha__year=periodo.objects.get(activo=True).anio).values_list('semana').distinct('semana'), fecha_inicio__year=periodo.objects.get(activo=True).anio), 'anio':periodo.objects.get(activo=True).anio}
	return render_to_response('general/eliminar-semana-calendario.html', ctx ,context_instance=RequestContext(request))

@permission_required('monitoreo.delete_calendario_academico', login_url='/inicio/')
def view_ajax_eliminar_semanacalendario(request, semana_id):
	if request.method == 'GET':
		try:
			sid = transaction.savepoint()
			calendario_academico.objects.filter(pk=semana_id).delete()
			transaction.savepoint_commit(sid)
			return HttpResponse(1)
		except Exception, e:
			print e
			transaction.savepoint_rollback(sid)
			return HttpResponse(0)
	return HttpResponse(0)